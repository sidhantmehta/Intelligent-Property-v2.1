#!/usr/bin/env python3
"""Converts a raw private-schools-register export (.xlsx) into the clean
CSV geo_model.private_schools imports (reference_data/private_schools_
greater_london.csv by default).

The source spreadsheet is scraped data, not a clean table: entries are
variable-length row blocks delimited by "Start"/"End" markers in column D,
each block holding label/value pairs (e.g. a "Gender Profile" row followed
by its value row). When a field is blank in the source, the scraper simply
omits it -- there's no blank placeholder row -- so a naive label-then-next-
row parse misreads the FOLLOWING label's text as the blank field's value
(e.g. Gender Profile's "value" comes back as the string "Size"). This
parser detects that case (the "value" it captured is itself one of the
known label strings, or a bare number where a string was expected) and
records the field as missing instead of misassigning it.

Usage:
    python scripts/convert_private_schools_xlsx.py INPUT.xlsx [--out PATH]
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl

LABELS = ["Gender Profile", "Size", "Day/boarding type", "Religious affiliation"]
LABELS_SET = set(LABELS)
FIELD_NAMES = ["name", "address", "postcode", "phone", "gender_profile", "size", "day_boarding_type", "religious_affiliation"]

_POSTCODE_RE = re.compile(r"([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})\s*$", re.IGNORECASE)


def _parse_block(block: list) -> dict:
    name = block[0]
    address = block[1] if len(block) > 1 else None
    phone = block[2] if len(block) > 2 else None

    fields: dict[str, object] = {}
    i = 3
    n = len(block)
    while i < n:
        val = block[i]
        if isinstance(val, str) and val in LABELS_SET and val not in fields:
            nxt = block[i + 1] if i + 1 < n else None
            if isinstance(nxt, str) and nxt in LABELS_SET:
                fields[val] = None  # blank -- the next label immediately follows
                i += 1
            elif val != "Size" and not isinstance(nxt, str):
                fields[val] = None  # non-Size field got a stray number -> value was blank
                i += 1
            else:
                fields[val] = nxt
                i += 2
        else:
            i += 1

    address = (address or "").strip()
    m = _POSTCODE_RE.search(address)
    postcode = m.group(1).upper() if m else None
    if postcode is None:
        raise ValueError(f"Could not extract a postcode from address {address!r} (school: {name!r})")

    return {
        "name": name,
        "address": address,
        "postcode": postcode,
        "phone": phone,
        "gender_profile": fields.get("Gender Profile"),
        "size": fields.get("Size"),
        "day_boarding_type": fields.get("Day/boarding type"),
        "religious_affiliation": fields.get("Religious affiliation"),
    }


def convert(input_path: Path, output_path: Path) -> int:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # row 1 is a stray header

    starts = [i for i, r in enumerate(rows) if r[3] == "Start"]
    ends = [i for i, r in enumerate(rows) if r[3] == "End"]
    if len(starts) != len(ends):
        raise ValueError(f"Mismatched Start/End markers: {len(starts)} starts, {len(ends)} ends")

    records = [_parse_block([rows[i][0] for i in range(s, e + 1)]) for s, e in zip(starts, ends)]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_NAMES)
        writer.writeheader()
        writer.writerows(records)

    return len(records)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", type=Path, help="Raw .xlsx export")
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "reference_data" / "private_schools_greater_london.csv",
    )
    args = parser.parse_args()
    n = convert(args.input, args.out)
    print(f"Wrote {n} schools to {args.out}")


if __name__ == "__main__":
    main()
